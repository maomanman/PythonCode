import os
import pandas as pd
import tkinter.messagebox as tkm
# from tkMessageBox import *
from tkinter.filedialog import *
import sys

import tkinter as tk

from tkinter import ttk
import re  # 用于字符串的多个分割符分割
from openpyxl import load_workbook # 不改变excel样式的修改excel文件

from openpyxl import load_workbook
from itertools import islice
import difflib # 用于模糊匹配

"""
功能说明：
用于自动分割：即根据已确定的索引段从原轨迹文件中截取出来，并创建新文件。
"""


class App:
    def __init__(self, root):
        # 创建一个框架，然后在里面添加一个Button按钮组件
        # (1)框架一般使用于在复杂的布局中起到将组建分组的作用
        select_file_frame = tk.Frame(root)
        select_file_frame.place(x=0, y=0, width=800, height=70)

        # (2)地块分段 y=50
        input_index_frame = tk.Frame(root)
        input_index_frame.place(x=10, y=70, width=300, height=230)
        # (3)道路分段  y=50
        road_frame = tk.Frame(root)
        road_frame.place(x=320, y=70, width=250, height=230)
        # (4)新文件命名标签 y=120
        new_file_name_frame = tk.Frame(root)
        new_file_name_frame.place(x=575, y=140, width=210, height=100)
        # (5)新文件的存放路径 y=290
        new_file_path_frame = tk.Frame(root)
        new_file_path_frame.place(x=0, y=310, width=800, height=40)
        # (6) 异常点移除：信号漂移点 y=330
        error_point_frame = tk.Frame(root)
        error_point_frame.place(x=0, y=350, width=800, height=30)

        # (7)确认分割按钮 y=370
        ok_group_frame = tk.Frame(root)
        ok_group_frame.place(x=0, y=390, width=800, height=50)

        # # # 创建一个按钮组件，fg是foreground的缩写，就是设置前景色的意思
        # (1)选择待分段为文件，并显示文件名
        # label1 = tk.Label(select_file_frame, text='目标类别选择:')
        # label1.place(x=20, y=15, width=80, height=13)
        # self.target_select = ttk.Combobox(select_file_frame, value=['汇总', '第一类别', '第二类别', '第三类别', '第四类别'])
        # self.target_select.current(0)
        # self.target_select.place(x=20, y=30, width=100, height=20)
        # self.target_select.bind("<<ComboboxSelected>>", self.targetSelect)



        label4 = tk.Label(select_file_frame, text='new index:')
        label4.place(x=130, y=15, width=80, height=13)
        self.new_file_index_text = tk.Entry(select_file_frame, font=4)
        self.new_file_index_text.place(x=140, y=30, width=60, height=20)

        # 读取新文件名的最大序列号
        self.wb = load_workbook(r'D:\mmm\轨迹数据集\轨迹索引-v1.0.xlsx')
        self.wb._active_sheet_index = 0  # 此属性是用来指定读取 excel 的页
        self.file_total_info_xlsx = self.wb.active  # 读取excel数据，默认读取第0页数据，_active_sheet_index指定页后，则读取指定页的数据
        self.file_name_index_max = int(
            self.file_total_info_xlsx['D' + str(self.file_total_info_xlsx.max_row)].value[0:5]) + 1
        self.new_file_index_text.insert(0, '{:0>5d}'.format(self.file_name_index_max))

        self.sheet_province = self.wb['Sheet4']
        self.sheet_city = self.wb['Sheet3']

        # 显示 省 市 县
        province = tk.Label(select_file_frame, text='省:')
        province.place(x=5, y=0, width=20, height=20)
        self.province_combobox = ttk.Combobox(select_file_frame)
        self.province_combobox.place(x=25, y=0, width=100, height=20)
        self.province_combobox['value'] = [self.sheet_province['A' + str(i)].value for i in
                                           range(2, self.sheet_province.max_row + 1)]
        city = tk.Label(select_file_frame, text='市:')
        city.place(x=5, y=23, width=20, height=20)
        self.city_combobox = ttk.Combobox(select_file_frame, value=['自动填写'])
        self.city_combobox.place(x=25, y=23, width=100, height=20)
        self.city_combobox.current(0)
        # self.city_combobox.bind("<<ComboboxSelected>>", self.targetSelect)
        county = tk.Label(select_file_frame, text='县:')
        county.place(x=5, y=46, width=20, height=20)
        self.county_combobox = ttk.Combobox(select_file_frame, value=['手动输入'])
        self.county_combobox.place(x=25, y=46, width=100, height=20)
        self.county_combobox.current(0)


        label2 = tk.Label(select_file_frame, text='old index:',fg='red')
        label2.place(x=220, y=15, width=60, height=13)
        self.old = StringVar()
        self.old_file_index_text = tk.Entry(select_file_frame, font=4, textvariable=self.old)
        self.old.set('001')
        self.old_file_index_text.place(x=220, y=30, width=38, height=20)

        # x=200 x=330 x=620
        self.select_file = tk.Button(select_file_frame, text="选择待分割文件", fg="blue", command=self.selectFile)
        self.select_file.place(x=300, y=15, width=100, height=30)
        self.show_filename1 = tk.Entry(select_file_frame, state='readonly', width=35)
        self.show_filename1.place(x=420, y=10, width=280, height=20)
        self.show_filename2 = tk.Entry(select_file_frame, state='readonly', width=35)
        self.show_filename2.place(x=420, y=31, width=280, height=20)
        self.select_end = tk.Button(select_file_frame, text="end", fg="blue", command=self.selectEnd)
        self.select_end.place(x=710, y=15, width=30, height=30)

        # (2)地块分段
        # input_index_frame = tk.Frame(root)
        # input_index_frame.place(x=10, y=50, width=300, height=230)
        # 显示需要分割的索引段，以及分段后的新文件名
        self.show_index_title = tk.Label(input_index_frame, text='地块索引段：', justify='left', font=6)
        self.show_index_title.place(x=0, y=0, width=140, height=20)
        self.show_index = tk.Text(input_index_frame)
        self.show_index.place(x=0, y=20, width=300, height=200)
        # self.input_ok = tk.Button(input_index_frame, text='确定索引段', font=6, command=self.inputIndexkOk)
        # self.input_ok.place(x=5, y=225, width=130, height=30)

        # (3)道路分段
        # road_frame = tk.Frame(root)
        # road_frame.place(x=320, y=50, width=250, height=240)
        self.road_index_title = tk.Label(road_frame, text='道路索引段：', justify='left', font=6)
        self.road_index_title.place(x=0, y=0, width=140, height=20)
        self.road_index = tk.Text(road_frame)
        self.road_index.place(x=0, y=20, width=250, height=200)
        # self.road_ok = tk.Button(road_frame, text='确定索引段', font=6, command=self.inputIndexkOk)
        # self.road_ok.place(x=60, y=225, width=130, height=30)

        # (4)新文件命名标签
        # new_file_name_frame = tk.Frame(root)
        # new_file_name_frame.place(x=575, y=70, width=210, height=100)
        self.type = tk.Label(new_file_name_frame, text='作业类型：', font=6, justify='left')
        self.type.place(x=0, y=0, width=115, height=20)
        self.type_select = ttk.Combobox(new_file_name_frame, value=['耕', '种', '收'])
        self.type_select.current(0)
        self.type_select.place(x=110, y=0, width=100, height=20)
        self.area = tk.Label(new_file_name_frame, text='面积尺度：', font=6)
        self.area.place(x=0, y=40, width=115, height=20)
        self.area_select = ttk.Combobox(new_file_name_frame, value=['大', '中', '小'])
        self.area_select.current(0)
        self.area_select.place(x=110, y=40, width=100, height=20)
        self.mode = tk.Label(new_file_name_frame, text='作业模式：', font=6)
        self.mode.place(x=0, y=80, width=115, height=20)
        self.mode_select = ttk.Combobox(new_file_name_frame, value=['套', '绕', '梭'])
        self.mode_select.current(0)
        self.mode_select.place(x=110, y=80, width=100, height=20)

        # (5)新文件的存放路径
        # new_file_path_frame = tk.Frame(root)
        # new_file_path_frame.place(x=0, y=290, width=800, height=40)
        self.new_path_button = tk.Button(new_file_path_frame, text="选择存放路径", fg="blue", command=self.selectNewPath)
        self.new_path_button.place(x=30, y=0, width=100, height=30)
        self.show_new_path = tk.Entry(new_file_path_frame)
        self.show_new_path.place(x=140, y=0, width=630, height=30)
        self.show_new_path.delete(0, END)
        self.show_new_path.insert(0, 'D:/mmm/轨迹数据集/汇总')

        # (6) 异常点移除：信号漂移点
        # error_point_frame = tk.Frame(root)
        # error_point_frame.place(x=0, y=330, width=800, height=30)
        label3 = tk.Label(error_point_frame, text='异常点：', font=('宋体', 10, 'bold'))
        label3.place(x=20, y=0, width=55, height=30)
        self.error_point_text = tk.Entry(error_point_frame, fg='red', font=('宋体', 12, 'bold'))
        self.error_point_text.place(x=80, y=0, width=700, height=30)

        # (7)确认分割按钮
        # ok_group_frame = tk.Frame(root)
        # ok_group_frame.place(x=0, y=370, width=800, height=50)
        self.input_batch = tk.Button(ok_group_frame, text='导入分段文件', font=6, command=self.inputIndexBatch)
        self.input_batch.place(x=100, y=0, width=140, height=50)
        self.patition_file_ok = tk.Button(ok_group_frame, text='确定分割文件', font=6, fg='red', command=self.patitionFileOk)
        self.patition_file_ok.place(x=560, y=0, width=140, height=50)

    def targetSelect(self, ty=0):
        """
        选择目标归类的类别
        ty 是 虚拟触发函数给的值
        :return:
        """
        print(self.city_combobox.current())
    #     cur_select = self.target_select.current()
    #
    #     if cur_select == 0:
    #         self.new_path = 'D:/mmm/轨迹数据集/汇总'
    #     elif cur_select == 1:
    #         self.new_path = 'D:/mmm/轨迹数据集/第1维度-按作业时间'
    #     elif cur_select == 2:
    #         self.new_path = 'D:/mmm/轨迹数据集/第2维度-按作业类型'
    #     elif cur_select == 3:
    #         self.new_path = 'D:/mmm/轨迹数据集/第3维度-按面积尺度'
    #     elif cur_select == 4:
    #         self.new_path = 'D:/mmm/轨迹数据集/第4维度-按作业模式'
    #     self.show_new_path.delete(0, END)
    #     self.show_new_path.insert(0, self.new_path)
        # self.file_name_info = pd.read_excel('D:/mmm/轨迹数据集/汇总/file_name_info.xlsx')
        # max_num = len(self.file_name_info) + 1
        # # self.new_file_index_text.delete(0, END)
        # self.new_file_index_text.insert(0, '{:0>5d}'.format(max_num))

    def selectFile(self):
        """
        选择待分割的文件
        :return:
        """
        f = askopenfilenames(title='选择一个待分割的文件', filetypes=[('*', 'csv'), ('*', 'xls'), ('*', 'xlsx')])
        if (f == '' or len(f) > 1):
            print("只能选择一个文件")
        else:
            f = f[0]
            self.filepath = os.path.split(f)[0]
            self.filename = os.path.split(f)[1]
            # print(self.filename)
            self.show_filename1['state'] = 'normal'
            self.show_filename1.delete(0, END)
            self.show_filename1.insert(0, self.filename)
            # self.show_filename1['state'] = 'disabled'
            if (self.filename.split('.')[1] == 'csv'):
                num = 7
            else:
                if (self.filename.split('.')[1] == 'xls'):
                    num = 7
                else:
                    num = 8
            self.show_filename2['state'] = 'normal'
            self.show_filename2.delete(0, END)
            self.show_filename2.insert(0, self.filename[0:-num])
            self.show_filename2['fg'] = 'tomato'
            # 索引段显示区需要重新可编辑
            self.show_index['state'] = 'normal'
            self.show_index.delete('0.0', END)
            self.show_index['fg'] = 'black'
            self.road_index['state'] = 'normal'
            self.road_index.delete('0.0', END)
            self.road_index['fg'] = 'black'

            # 清空异常输入框
            self.error_point_text.delete(0, END)

            # 显示 省市县
            # self.province_combobox
            sheet_provinces = list(self.sheet_province.values)
            sheet_provinces = [sheet_provinces[i][0] for i in range(0, len(sheet_provinces))]

            sheet_citys = self.sheet_city.values
            cols = next(sheet_citys)[1:]
            sheet_citys = list(sheet_citys)
            data = (islice(r, 1, None) for r in sheet_citys)
            sheet_citys = pd.DataFrame(data, columns=cols)

            strPath=self.filepath.split('/')
            province = difflib.get_close_matches(strPath[-1], sheet_provinces, 1, cutoff=0.5)
            if  province==[]: # strPath[-1]  not in sheet_provinces: # 省
                city = difflib.get_close_matches(strPath[-1], sheet_citys[:].values, 1, cutoff=0.5)
                if strPath[-1]  in sheet_citys[:].values: # 市
                    self.city_combobox.delete(0,END)
                    self.city_combobox.insert(0,strPath[-1]) # 倒数 第一个是 市集
                    self.province_combobox.delete(0,END)
                    self.province_combobox.insert(0,strPath[-2]) # 倒数 第二个是 省级
                    self.county_combobox.delete(0,END) # 没有县级
                else:
                    self.county_combobox.delete(0, END)
                    self.county_combobox.insert(0, strPath[-1]) # 倒数 第一个是 县集
                    self.city_combobox.delete(0, END)
                    self.city_combobox.insert(0, strPath[-2])  # 倒数 第二个是 市集
                    self.province_combobox.delete(0, END)
                    self.province_combobox.insert(0, strPath[-3]) # 倒数 第三个是 省集
            else:
                self.province_combobox.delete(0,END)
                self.province_combobox.insert(0,province[0])
                self.county_combobox.delete(0, END)
                self.city_combobox.delete(0, END)


    def selectEnd(self):
        """
        对于源文件名后没有类型说明 但是在源文件选择时 被截取了 类型说明所占字段  而出现的错误进行纠错
        :return:
        """
        if (self.filename.split('.')[1] == 'csv'):
            num = 4
        else:
            if (self.filename.split('.')[1] == 'xls'):
                num = 4
            else:
                num = 5
        self.show_filename2['state'] = 'normal'
        self.show_filename2.delete(0, END)
        self.show_filename2.insert(0, self.filename[0:-num])
        self.show_filename2['fg'] = 'tomato'

    def selectNewPath(self):
        self.new_path = askdirectory()
        if self.new_path:
            self.show_new_path['state'] = 'normal'
            self.show_new_path.delete(0, END)
            self.show_new_path.insert(0, self.new_path)
            # self.show_new_path['state'] = 'disabled'

    def inputIndexBatch(self):
        """
        文件导入索引段，并显示在左侧列表栏
        :return:
        """
        # print(type(self.type_select.get()))
        # self.show_index.insert("1", '这是一个测试\n')
        print(self.city_combobox.current())
        print(self.city_combobox.get())
        self.county_combobox.delete(0,END)
        self.county_combobox.insert(0,'自动输入')
        # todo 完成导入分段文件的函数处理

    def patitionFileOk(self):
        """
        确定分割文件
        :return:
        """
        self.show_index['fg'] = 'green'
        self.show_index['state'] = 'disabled'
        self.road_index['fg'] = 'green'
        self.road_index['state'] = 'disabled'
        # 索引号都是字符串类型
        filedIndex = self.show_index.get("0.0", "end")
        roadIndex = self.road_index.get("0.0", "end")

        filedIndex = self._str2list(filedIndex)
        roadIndex = self._str2list(roadIndex)

        newFiledFileName = self._patitionFile(filedIndex)
        newRoadFileName = self._patitionFile(roadIndex, flag=0)

        # 将旧文件进行重命名
        old_index = self.old_file_index_text.get()
        os.rename(self.filepath + '/' + self.filename, self.filepath + '/' + old_index + '-' + self.filename)
        self.old.set('{:0>3d}'.format(int(old_index) + 1))
        # self.file_total_info_xlsx.save()
        self.wb.save(r'D:\mmm\轨迹数据集\轨迹索引-v1.0.xlsx')
        # self.file_name_info.to_excel(self.new_path+'\\file_name_info.xlsx',index=False)

        tkm.showinfo('提示', '分割成功')

    def _str2list(self, s):
        """
        将text输入的字符串转换成list，便于分割
        :param s: 待转换的字符串
        :return: list 易于分割的列表
        """
        l = s.split('\n')
        while '' in l:  # 删除空元素
            l.remove('')
        count = len(l)  # 分割后的文件个数
        for i in range(0, count):
            p = l[i].count(':') + l[i].count('：')
            if p > 1:
                l[i] = re.split(';|；', l[i])

        return l

    def _patitionFile(self, index, flag=1):
        """
        根据索引段分割文件
        :param index:
        :param flag = 1 表示分割的是地块文件； = 0 表示分割的道路文件
        :return: 新文件名字符串
        """
        # 读取待分割轨迹文件

        if (self.filename.split('.')[1] == 'csv'):
            data = pd.read_csv(self.filepath + '/' + self.filename)
        else:
            data = pd.read_excel(self.filepath + '/' + self.filename)

        newFileName = []  # 新文件名集合，用于最后显示的

        self.new_path = self.show_new_path.get()
        self.filename2 = self.show_filename2.get()
        # 获取界面的异常点集
        if self.error_point_text.get():
            error_point = [int(i) for i in re.split(' |,|，', self.error_point_text.get())]
        else:
            error_point = []

        for i in range(0, len(index)):
            newData = pd.DataFrame(columns=data.columns)
            if type(index[i]) == list:  # 一个地块由多个索引段组成
                for j in range(0, len(index[i])):
                    startIndex = int(re.split(':|：', index[i][j])[0]) - 1
                    endIndex = int(re.split(':|：', index[i][j])[1]) - 1
                    newData = newData.append(data.loc[startIndex:endIndex])

            else:
                startIndex = int(re.split(':|：', index[i])[0]) - 1
                endIndex = int(re.split(':|：', index[i])[1]) - 1
                newData = newData.append(data.loc[startIndex:endIndex])

            try:
                newData.reset_index(drop=True, inplace=True)
                tm = newData.loc[0, 'GPS时间']
                if type(tm) == str:
                    tm = pd.Timestamp(tm)
                date = '{:0>2d}'.format(tm.date().month) + '{:0>2d}'.format(tm.date().day)
                time = '{:0>2d}'.format(tm.time().hour) + '{:0>2d}'.format(tm.time().minute)
                if flag == 1:  # 地块轨迹文件名
                    new_index = self.new_file_index_text.get()
                    name = new_index + ' ' + self.type_select.get() + '-' + self.area_select.get() + '-' + self.mode_select.get() + '==' + self.filename2 + '==' + date + '-' + time + '-filed'
                    self.new_file_index_text.delete(0,END)
                    self.new_file_index_text.insert(0,'{:0>5d}'.format(int(new_index) + 1))
                    #将新文件名登记到 轨迹索引表中
                    self.file_total_info_xlsx['D'+str((int(new_index) + 2))]=name+'.xlsx'   # new_index 指excel新一行文件名的序号， 2 是文件名序号与表行号的差
                    # 添加 将文件名 拆解的函数
                    excelE = '=LEFT(D{},5)'.format(str((int(new_index) + 2)))
                    self.file_total_info_xlsx['E' + str((int(new_index) + 2))]=excelE
                    excelF='=MID(D{},7,1)'.format(str((int(new_index) + 2)))
                    self.file_total_info_xlsx['F' + str((int(new_index) + 2))] = excelF
                    excelG='=SWITCH(MID(D{},9,1),"大","大块面积","中","中等面积","小","小块面积","-","计算错误","=","计算错误")'.format(str((int(new_index) + 2)))
                    self.file_total_info_xlsx['G' + str((int(new_index) + 2))] = excelG
                    excelH='=SWITCH(MID(D{},11,1),"套","套行法","绕","绕行法","梭","梭行法","-","计算错误","=","计算错误")'.format(str((int(new_index) + 2)))
                    self.file_total_info_xlsx['H' + str((int(new_index) + 2))] = excelH
                    self.file_total_info_xlsx['A' + str((int(new_index) + 2))]=self.province_combobox.get()
                    self.file_total_info_xlsx['B' + str((int(new_index) + 2))] = self.city_combobox.get()
                    self.file_total_info_xlsx['C' + str((int(new_index) + 2))] = self.county_combobox.get()


                elif flag == 0:
                    name = self.filename2 + '==' + date + '-' + time + '-road'


                # 新文件名加入文件名集合
                newFileName.append(name)
                # 更改新文件的索引号
                newData.set_index("序列号", drop=True, inplace=True)

                # 删除新文件中信号漂移点
                for ep in error_point:
                    if ep in newData.index:
                        newData.drop(ep, axis=0, inplace=True)  # 删除行

                if flag == 1:  # 地块轨迹文件名:
                    newData.to_excel(self.new_path + '\\' + name + '.xlsx')
                else:
                    newData.to_excel(r'D:\mmm\轨迹数据集\汇总\road\\' + name + '.xlsx')
            except Exception as e:
                # print(tkinter.messagebox.showerror("Error", " " + repr(e) + os.linesep))
                tkm.showerror('分割错误', repr(e))
            del newData
            del name

        # 将新文件名进行显示
        if flag == 1:  # 地块轨迹文件名
            self.show_index['state'] = 'normal'
            self.show_index.insert('insert', '\n\n' + '\n'.join(newFileName))
            self.show_index['state'] = 'disable'

        else:
            self.road_index['state'] = 'normal'
            self.road_index.insert('insert', '\n\n' + '\n'.join(newFileName))
            self.road_index['state'] = 'disable'

    def readDir(self,file_dir,type='xlsx'):
        """
        读取目录下 指定文件类型的文件个数，及文件名，
        :param path: 待读取路径
        :param type: 待读取的文件类型
        :return:
        """
        file_info = pd.DataFrame(columns=['filename'])
        fn=os.listdir(file_dir)
        count = len(fn)
        i=0
        for file_cur  in fn:
            path = os.path.join(file_dir,file_cur)
            if os.path.isfile(path):
                if file_cur!='file_name_info.xlsx':
                    file_info.loc[i]=[file_cur]
                    i=i+1
        file_info.to_excel(file_dir+'\\file_name_info.xlsx',index=False)


# 创建一个toplevel的根窗口，并把他作为擦参数实例化APP对象
root = tk.Tk()
root.title('根据序列号分割文件')
root.geometry("800x450+500+200")


app = App(root)
# 开始主事件循环
root.mainloop()
del app
del root
