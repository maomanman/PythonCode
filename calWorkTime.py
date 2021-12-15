import  pandas as pda
from openpyxl import load_workbook
import math
import  sys
import os
# sys.path.append("D:/mmm/python/PythonCode/alpha_shape")
# sys.path.insert(0, os.path.abspath('..'))
#
# sys.path.insert(0, os.path.abspath('.'))
from alpha_shape.ZuoBiaoZhuanHuan import LatLon2GSXY

def calWorkTime(data):
    """
    计算一块地的总作业时间
    :param filename: 带路径的轨迹点文件
    :return: 作业总时间   和 幅宽
    """

        # info=pda.DataFrame(columns=['filename','width'])

    # print(filename)
    # data = pda.read_excel(filename) # 打开轨迹点文件

    count = data.shape[0]  # 获取轨迹点文件总行数
    # 通过序列号的连续行计算工作时长 begin
    # flag = 0 # 标识开始结束时间
    # sumTime = pda.to_timedelta(0)
    # for i in range(0, count - 1):  # range 包含 起点 不包含 终点
    #     if flag == 0:
    #         startTime = data.loc[i, 'GPS时间']
    #         print('start：{}  {}'.format(data.loc[i, '序列号'],startTime))
    #         flag = 1
    #     if data.loc[i + 1, '序列号'] - data.loc[i, '序列号'] == 1:  # 说明轨迹点不连续
    #         if i == count - 2:
    #             endTime = data.loc[i + 1, 'GPS时间']
    #             print('end：{} {}'.format(data.loc[i+1, '序列号'],endTime))
    #         else:
    #             continue
    #     else:
    #         endTime = data.loc[i, 'GPS时间']
    #         print('end：{} {}'.format(data.loc[i, '序列号'],endTime))
    #
    #     if type(startTime) == str:
    #         startTime = pda.Timestamp(startTime)
    #         endTime = pda.Timestamp(endTime)
    #     sumTime = sumTime + (endTime - startTime)
    #     print(sumTime)
    #     flag = 0
    # 通过序列号的连续行计算工作时长 end

    # print((float(data.loc[:, '幅宽(m)'].mode())))

    sumTime = pda.to_timedelta(0)
    # 通过采样时间计算工作时长 begin

    sumTime = pda.to_timedelta(0)
    for i in range(1,count):
        if type(data.loc[i,'GPS时间']) == str:
            yangT= pda.Timestamp(data.loc[i,'GPS时间'])-pda.Timestamp(data.loc[i-1,'GPS时间'])
        else:
            yangT =data.loc[i, 'GPS时间'] - data.loc[i - 1, 'GPS时间']
        if yangT.total_seconds() < 60:  # 间隔时间小于1分钟的记录为工作时间，间隔时间大于1分钟的则不记录为工作时间
            sumTime = sumTime + yangT
        else: # 时间间隔大于1分钟的计算两点间距离
            x1, y1= LatLon2GSXY(data.loc[i-1, '纬度'], data.loc[i-1, '经度'])
            x2, y2 = LatLon2GSXY(data.loc[i, '纬度'], data.loc[i, '经度'])
            d = math.sqrt(pow(y2-y1,2)+pow(x2-x1,2))
            v = d/yangT.total_seconds()   # m/s

            if v > 1 and v < 2.7778 : # 平均速度 >3.6km/h  <10km/h
                sumTime = sumTime + yangT
            # print('前：{}'.format(data.loc[i-1, '速度(km/h)']*0.2777778))
            # else:
            #     print('v：{}'.format(v))
            # print('后：{}'.format(data.loc[i, '速度(km/h)'] * 0.2777778))

    # 通过采样时间计算工作时长 end
    # print(data.loc[:, 'time'].nunique())
    return sumTime,float(data.loc[:, '幅宽(m)'].mode())

if __name__ == '__main__':
    path = 'D:\mmm\轨迹数据集\轨迹索引-v1.0.xlsx'
    wb = load_workbook(path)
    wb._active_sheet_index = 0  # 此属性是用来指定读取 excel 的页
    ws = wb.active  # 读取excel数据，默认读取第0页数据，_active_sheet_index指定页后，则读取指定页的数据
    # ps=wb['汇总']
    # data = ps.values
    # cols = next(data)[1:]
    # data = list(data)

    for i in range(3,ws.max_row+1):
        name = ws['D'+str(i)].value
        filename = 'D:\mmm\轨迹数据集\汇总\\'+name
        data = pda.read_excel(filename) # 打开轨迹点文件
        # worktime,width = calWorkTime(data)
        # ws['L' + str(i)]=str(worktime) # 将工作时长写入索引文件
        ws['O' + str(i)] = data.shape[0] # 将幅宽写汝索引文件

    # worktime,width=calWorkTime('D:\mmm\轨迹数据集\汇总\\'+ws['D53'].value)
    # ws['N53' ] = str(worktime) # 将工作时长写入索引文件
    # ws['O53'] = width  # 将幅宽写汝索引文件
    # print( worktime,width)
    wb.save(path)
    # print(ws.max_row,ws['D39'].value)
