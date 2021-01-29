import os
import pandas as pd
import time
from itertools import islice

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import difflib


def readDir(file_dir, type='xlsx'):
    """
    读取目录下 指定文件类型的文件个数，及文件名，
    :param path: 待读取路径
    :param type: 待读取的文件类型
    :return:
    """
    file_info = pd.DataFrame(columns=['filename'])
    fn = os.listdir(file_dir)
    count = len(fn)
    i = 0
    for file_cur in fn:
        path = os.path.join(file_dir, file_cur)
        if os.path.isfile(path):
            if file_cur != 'file_name_info.xlsx':
                file_info.loc[i] = [file_cur]
                i = i + 1
    file_info.to_excel(file_dir + '\\file_name_info.xlsx', index=False)


def readExcel():
    """
    读取 xlsx 文件，不改变原有文件样式 写入 数据
    :return:
    """

    wb = load_workbook(r'D:\mmm\轨迹数据集\轨迹索引-v1.0 - 副本.xlsx')
    wb._active_sheet_index = 0  # 此属性是用来指定读取 excel 的页
    ws = wb.active  # 读取excel数据，默认读取第0页数据，_active_sheet_index指定页后，则读取指定页的数据
    # ws['D57'] = '测试一下不改变格式'
    ps=wb['Sheet3']
    data = ps.values
    cols = next(data)[1:]
    data = list(data)
    data = (islice(r, 1, None) for r in data)
    df = pd.DataFrame(data,  columns=cols)

    # data = list(ps.values)
    # data = [data[i][0]  for i in  range(0,len(data))]
    # print('甘肃省' in data)
    # print(data)
    print('武汉' in df[:].values)
    print('武汉市' in df[:].values)
    dd=difflib.get_close_matches('武汉市', df[:].values, 1, cutoff=0.5)
    # print('天津市' in df[:].values)
    print(dd)
    print(dd == [])

    # num = ws['D'].
    # data=[ws['D' + str(i)].value for i in range(1, ws.max_row) ]
    # print(ws.max_row)
    # print('[{}]'.format(ws['D' + str(ws.max_row)].value[0:5]))
    # print('E55:',ws['E55'].value)
    # ws['E56']=ws.cell(row=55,column=5).value
    # print('E56:',ws['E56'].value)
    # for i in range(1, ws.max_row):
    #     print('D' + str(i))
    #     print(ws['D' + str(i)].value)
    #     data.append(ws['D' + str(i)].value )
    # data = ws.values
    # cols = next(data)[1:]
    # data = list(data)
    # idx = [r[0] for r in data]
    # data = (slice(r, 1, None) for r in data)\
    # print(data)
    # df = pd.DataFrame(data)
    # print(df.isnull())
    # for flag in df.isnull():
    #     if flag:
    #         df.dropna(axis=1,how='any',inplace=True)
    # df.dropna(axis=0, how='any', inplace=True)
    # print(df.shape)
    # df.to_excel(r'D:\mmm\轨迹数据集\444.xlsx',index=False)
    # # data = ws['D']
    # print(ws.max_row)
    # print(ws.values)
    # for r in dataframe_to_rows(df, index=True, header=True):
    #     ws.append(r)
    # wb.save(r'D:\mmm\轨迹数据集\轨迹索引-v1.0 - 副本.xlsx')


# start = time.time()
# readDir(r'D:\mmm\轨迹数据集\汇总')
# end =time.time()
# print(end-start)

readExcel()
